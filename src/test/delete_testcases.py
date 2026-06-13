"""
DeleteTestCases — reads test case keys from resources/delete_Test_case.xlsx
and deletes them from Jira Zephyr Scale.

Expected Excel columns:
    Test_case  : key eg.Test Case ID  (required)
    Folder     : Folder path (required)

Usage:
    python -m test.delete_testcases
    python -m test.delete_testcases --dry-run
    python -m test.delete_testcases --file resources/delete_Test_case.xlsx
"""

import sys
import os
import argparse
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from main.config import AppConfig
from main.jira_client import JiraClient
from main.report import ReportGenerator


class DeleteTestCases:

    DEFAULT_XLSX = os.path.join(
        os.path.dirname(__file__), "..", "..", "resources", "delete_Test_case.xlsx"
    )

    def __init__(self, xlsx_path: str = "", dry_run: bool = False):
        self._cfg       = AppConfig()
        self._client    = JiraClient()
        self.xlsx_path  = xlsx_path or self.DEFAULT_XLSX
        self.dry_run    = dry_run

    # — Read ————————————————————————————————————————————————————————

    def load_xlsx(self) -> list:
        """Read Excel and return list of {key, folder} dicts."""
        if not os.path.exists(self.xlsx_path):
            print(f"[ERROR] Excel file not found: {os.path.abspath(self.xlsx_path)}")
            sys.exit(1)

        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_lookup = {h.lower(): idx for idx, h in enumerate(raw_headers)}

        key_idx    = next((header_lookup.get(c) for c in ("test_case", "test case", "key") if header_lookup.get(c) is not None), None)
        folder_idx = header_lookup.get("folder")
        name_idx   = header_lookup.get("name")

        if key_idx is None:
            print(f"[ERROR] Could not find a 'Test_case' or 'Key' column. Headers: {raw_headers}")
            sys.exit(1)

        if folder_idx is None:
            print(f"[ERROR] 'Folder' column is required but not found. Headers: {raw_headers}")
            sys.exit(1)

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = str(row[key_idx]).strip() if row[key_idx] is not None else ""
            if not key or key == "None":
                continue
            name   = str(row[name_idx]).strip() if name_idx is not None and row[name_idx] else ""
            folder = str(row[folder_idx]).strip() if row[folder_idx] else ""
            rows.append({"key": key, "name": name or key, "folder": folder})

        wb.close()
        print(f"  Loaded {len(rows)} test case(s) to delete from: {os.path.basename(self.xlsx_path)}\n")
        return rows

    # — Delete ————————————————————————————————————————————————————————

    def delete_test_case(self, key: str) -> tuple:
        """DELETE /testcase/{key}. Returns (success, message)."""
        status, body = self._client.delete(f"/testcase/{key}")
        if status in (200, 204):
            return True, "deleted"
        return False, f"{status} — {body}"

    # — Run ————————————————————————————————————————————————————————————

    def run(self):
        print("=" * 65)
        print("  Jira Zephyr Scale — Delete Test Cases from Excel")
        print("=" * 65)
        print(f"  Project  : {self._cfg.project_key}")
        print(f"  Excel File: {os.path.abspath(self.xlsx_path)}")
        if self.dry_run:
            print("  Mode     : DRY RUN (no changes will be made)")
        print("=" * 65 + "\n")

        records = self.load_xlsx()

        if not records:
            print("[WARN] No valid rows found in Excel file.")
            sys.exit(0)

        ok_list, fail_list = [], []

        for i, record in enumerate(records, 1):
            key = record["key"]
            print(f"  [{i:>3}/{len(records)}] {key:20s} ... ", end="", flush=True)

            if self.dry_run:
                print("SKIP (dry-run)")
                continue

            success, msg = self.delete_test_case(key)
            if success:
                print(f"✅  Deleted")
                ok_list.append(key)
            else:
                print(f"❌  {msg}")
                fail_list.append(key)

        print()
        print("=" * 65)
        if self.dry_run:
            print(f"  DRY RUN — {len(records)} test case(s) would be deleted.")
        else:
            print(f"  SUMMARY")
            print(f"  Deleted : {len(ok_list)}")
            print(f"  Failed  : {len(fail_list)}")
            if ok_list:
                print(f"\n  Deleted keys: {', '.join(ok_list)}")
            if fail_list:
                print(f"\n  Failed keys:")
                for k in fail_list:
                    print(f"    — {k}")
        print("=" * 65)

        if not self.dry_run:
            reports_dir = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
            os.makedirs(reports_dir, exist_ok=True)
            report_path = os.path.join(reports_dir, "Delete_TC_Report.html")
            test_cases  = [{"key": r["key"], "name": r.get("name", r["key"])} for r in records]
            folder_path = records[0].get("folder", "") if records else ""
            ReportGenerator(output_path=report_path, action_label="Deleted", report_title="Zephyr Scale — Delete Test Cases Report").generate(
                test_cases=test_cases,
                ok_list=ok_list,
                fail_list=fail_list,
                project_key=self._cfg.project_key,
                folder_path=folder_path,
            )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Delete test cases from Jira Zephyr Scale using an Excel file")
    parser.add_argument("--file",    default="", help="Path to Excel file (default: resources/delete_Test_case.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without deleting")
    args = parser.parse_args()
    DeleteTestCases(xlsx_path=args.file, dry_run=args.dry_run).run()
