"""
LinkTestCasesToIssue — reads (Key, Traceability) pairs from
resources/update_test_case.xlsx and adds each test case as a traceability
(issue link) on the corresponding Jira issue via the Zephyr Scale ATM API.

Zephyr Scale endpoint used:
    PUT /rest/atm/1.0/testcase/{testCaseKey}
    Body: { "issueLinks": ["ISSUE-KEY", ...] }

The script merges the supplied issue keys with any already present on the
test case so that existing links are never removed.

Expected Excel columns (from update_test_case.xlsx):
    Key          : Zephyr test case key  (e.g. PROJECT-T1234)
    Traceability : Jira issue key to link TO (e.g. PROJECT-5678)

Rows with an empty Key or empty Traceability are skipped.

Usage:
    python -m test.link_testcases_to_issue
    python -m test.link_testcases_to_issue --dry-run
    python -m test.link_testcases_to_issue --file resources/update_test_case.xlsx
"""

import sys
import os
import json
import argparse

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from main.config import AppConfig
from main.jira_client import JiraClient
from main.report import ReportGenerator


class LinkTestCasesToIssue:

    DEFAULT_XLSX = os.path.join(
        os.path.dirname(__file__), "..", "..", "resources", "update_test_case.xlsx"
    )

    def __init__(self, xlsx_path: str = "", dry_run: bool = False):
        self._cfg       = AppConfig()
        self._client    = JiraClient()
        self.xlsx_path  = xlsx_path or self.DEFAULT_XLSX
        self.dry_run    = dry_run

    # — Read ————————————————————————————————————————————————————————

    def load_xlsx(self) -> list:
        """Read Excel and return list of {tc_key, issue_key} dicts."""
        if not os.path.exists(self.xlsx_path):
            print(f"[ERROR] Excel file not found: {os.path.abspath(self.xlsx_path)}")
            sys.exit(1)

        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        header_lookup = {h.lower(): idx for idx, h in enumerate(raw_headers)}

        tc_idx    = header_lookup.get("key")
        issue_idx = header_lookup.get("traceability")

        if tc_idx is None:
            print(f"[ERROR] 'Key' column not found. Headers: {raw_headers}")
            sys.exit(1)
        if issue_idx is None:
            print(f"[ERROR] 'Traceability' column not found. Headers: {raw_headers}")
            sys.exit(1)

        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            tc_key    = str(row[tc_idx]).strip() if row[tc_idx] is not None else ""
            issue_key = str(row[issue_idx]).strip() if row[issue_idx] is not None else ""
            if not tc_key or tc_key == "None":
                continue
            if not issue_key or issue_key == "None":
                continue
            rows.append({"tc_key": tc_key, "issue_key": issue_key})

        wb.close()
        print(f"  Loaded {len(rows)} link(s) from: {os.path.basename(self.xlsx_path)}\n")
        return rows

    # — Group rows by test case ————————————————————————————————————————

    def group_by_tc(self, rows: list) -> dict:
        """Merge multiple issue keys that target the same test case."""
        grouped = {}
        for r in rows:
            grouped.setdefault(r["tc_key"], set()).add(r["issue_key"])
        return grouped

    # — Fetch existing test case ————————————————————————————————————————

    def fetch_existing(self, tc_key: str) -> tuple:
        """GET /testcase/{key}. Returns (status_code, data_dict)."""
        return self._client.get(f"/testcase/{tc_key}")

    # — Link ————————————————————————————————————————————————————————————

    def link_test_case(self, tc_key: str, issue_keys: set) -> tuple:
        """PUT merged issueLinks onto the test case, preserving all existing fields.
        Returns (success, message).
        """
        get_status, existing = self.fetch_existing(tc_key)

        if get_status == 404:
            return False, f"Test case '{tc_key}' not found in Zephyr (404)"
        if get_status == 401:
            return False, "Unauthorized — check jira_token in config.properties (401)"
        if get_status != 200:
            return False, f"Could not fetch test case '{tc_key}': HTTP {get_status}"

        # Merge new issue keys with any already on the test case
        merged_links = sorted(set(existing.get("issueLinks", [])) | issue_keys)

        # Build a minimal payload that satisfies Zephyr's required-field validation
        # by echoing back the fields already stored on the test case
        payload = {"issueLinks": merged_links}

        for field in ("folder", "status", "priority", "owner"):
            if existing.get(field):
                payload[field] = existing[field]

        if existing.get("customFields"):
            payload["customFields"] = existing["customFields"]

        put_status, body = self._client.put(f"/testcase/{tc_key}", payload)
        if put_status in (200, 204):
            return True, ", ".join(sorted(issue_keys))
        try:
            detail = json.loads(body) if isinstance(body, str) else body
            msg    = detail.get("message") or detail.get("error") or str(body)[:200]
        except Exception:
            msg = str(body)[:200]
        return False, f"{put_status} — {msg}"

    # — Run ————————————————————————————————————————————————————————————

    def run(self):
        print("=" * 65)
        print("  Jira Zephyr Scale — Link Test Cases to Jira Issues")
        print("=" * 65)
        print(f"  Project  : {self._cfg.project_key}")
        print(f"  Excel File: {os.path.abspath(self.xlsx_path)}")
        if self.dry_run:
            print("  Mode     : DRY RUN (no changes will be made)")
        print("=" * 65 + "\n")

        rows    = self.load_xlsx()
        if not rows:
            print("[WARN] No valid rows found in Excel file.")
            sys.exit(0)

        grouped   = self.group_by_tc(rows)
        ok_list   = []
        fail_list = []
        test_cases = []

        total = len(grouped)
        for i, (tc_key, issue_keys) in enumerate(grouped.items(), 1):
            label = f"{tc_key} → {', '.join(sorted(issue_keys))}"
            print(f"  [{i:>3}/{total}] {label:<58s} ... ", end="", flush=True)

            if self.dry_run:
                print("SKIP (dry-run)")
                test_cases.append({"key": tc_key, "name": ", ".join(sorted(issue_keys))})
                continue

            success, msg = self.link_test_case(tc_key, issue_keys)
            if success:
                print(f"✅  Linked")
                ok_list.append(tc_key)
            else:
                print(f"❌  {msg}")
                fail_list.append(tc_key)
            test_cases.append({"key": tc_key, "name": ", ".join(sorted(issue_keys))})

        print()
        print("=" * 65)
        if self.dry_run:
            print(f"  DRY RUN — {total} test case(s) would be linked.")
        else:
            print(f"  SUMMARY")
            print(f"  Linked  : {len(ok_list)}")
            print(f"  Failed  : {len(fail_list)}")
            if ok_list:
                print(f"\n  Linked keys: {', '.join(ok_list)}")
            if fail_list:
                print(f"\n  Failed:")
                for k in fail_list:
                    print(f"    — {k}")
        print("=" * 65)

        if not self.dry_run:
            cfg         = self._cfg
            reports_dir = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
            os.makedirs(reports_dir, exist_ok=True)
            report_path = os.path.join(reports_dir, "Link_TC_Report.html")
            ReportGenerator(
                output_path=report_path,
                action_label="Linked",
                report_title="Zephyr Scale — Link Test Cases to Issues Report",
            ).generate(
                test_cases=test_cases,
                ok_list=ok_list,
                fail_list=fail_list,
                project_key=cfg.project_key,
                folder_path="",
            )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Link test cases to Jira issues via Zephyr Scale"
    )
    parser.add_argument("--file",    default="", help="Path to Excel file (default: resources/update_test_case.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without linking")
    args = parser.parse_args()
    LinkTestCasesToIssue(xlsx_path=args.file, dry_run=args.dry_run).run()
