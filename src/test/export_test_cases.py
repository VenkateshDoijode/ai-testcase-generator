"""
CycleExporter — exports Jira Zephyr Scale test cycle executions to Excel (.xlsx).

Usage:
    python -m test.export_test_cycle
    python -m test.export_test_cycle --cycle <CYCLE_KEY>
    python -m test.export_test_cycle --cycle <CYCLE_KEY> --out results.xlsx
"""

import sys
import os
import argparse

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from main.config import AppConfig
from main.jira_client import JiraClient


class CycleExporter:

    COLUMNS = [
        "Key",
        "Test Case Name",
        "Status",
        "Environment",
        "Assigned To",
        "Executed By",
    ]

    def __init__(self, cycle_key: str = "", out_path: str = ""):
        self._cfg       = AppConfig()
        self._client    = JiraClient()
        self.cycle_key  = (cycle_key or self._cfg.test_cycle).strip()
        if not self.cycle_key:
            print("[ERROR] test_cycle is not set in config.properties under [testcase] section.")
            print("        Add: test_cycle = <CYCLE_KEY>")
            sys.exit(1)
        self._reports_dir = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
        os.makedirs(self._reports_dir, exist_ok=True)
        self._out_path_override = out_path

    # — Fetch ————————————————————————————————————————————————————————

    def fetch_cycle(self) -> dict:
        """GET /testrun/{key} and return the cycle dict."""
        status, data = self._client.get(f"/testrun/{self.cycle_key}")
        if status == 200:
            return data
        if status == 404:
            print(f"[ERROR] Cycle '{self.cycle_key}' not found. Check the cycle key in config.properties.")
        elif status == 401:
            print(f"[ERROR] Unauthorized — check jira_token in config.properties.")
        else:
            print(f"[ERROR] Failed to fetch cycle '{self.cycle_key}': HTTP {status} — {data}")
        sys.exit(1)

    def fetch_executions(self, cycle: dict) -> list:
        """Return the executions embedded in the cycle's `items` list."""
        return cycle.get("items", [])

    def fetch_tc_names(self, keys: list) -> dict:
        """Fetch test case names for a list of keys. Returns {key: name}."""
        names = {}
        for key in keys:
            status, data = self._client.get(f"/testcase/{key}")
            if status == 200:
                names[key] = data.get("name", key)
            else:
                names[key] = key
        return names

    # — Export ————————————————————————————————————————————————————————

    def write_excel(self, executions: list, tc_names: dict = None):
        """Write execution results to an Excel (.xlsx) file."""
        tc_names = tc_names or {}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cycle Executions"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill(fill_type="solid", fgColor="0052CC")
        header_align = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(self.COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_align

        for row_idx, ex in enumerate(executions, 2):
            for col_idx, value in enumerate(self._row(ex, tc_names), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(self.out_path)
        print(f"\n  ✅  Excel saved → {os.path.abspath(self.out_path)}")

    def _row(self, ex: dict, tc_names: dict = None) -> list:
        tc_names = tc_names or {}
        tc_key   = ex.get("testCaseKey", "")
        tc_name  = tc_names.get(tc_key, "")
        status   = ex.get("status", "")
        env      = ex.get("environment", "")
        assigned = self._extract_user(ex.get("assignedTo", ""))
        executed = self._extract_user(ex.get("executedBy", ""))
        return [tc_key, tc_name, status, env, assigned, executed]

    @staticmethod
    def _extract_user(val) -> str:
        if not val:
            return ""
        if isinstance(val, dict):
            return val.get("displayName") or val.get("name") or val.get("accountId") or ""
        return str(val)

    # — Run ————————————————————————————————————————————————————————————

    def run(self):
        print("=" * 60)
        print("  Zephyr Scale — Export Cycle Executions to Excel")
        print("=" * 60)
        print(f"  Cycle  : {self.cycle_key}")
        print("=" * 60 + "\n")

        print("Fetching cycle...")
        cycle      = self.fetch_cycle()
        cycle_name = cycle.get("name", self.cycle_key).strip()
        safe_name  = self.cycle_key.replace("/", "_")
        self.out_path = self._out_path_override or os.path.join(self._reports_dir, f"{safe_name}.xlsx")
        summary    = cycle.get("executionSummary", {})
        print(f"  Cycle key : {cycle.get('key', self.cycle_key)}")
        print(f"  Cycle name: {cycle_name}")
        print(f"  Version   : {cycle.get('version', 'N/A')}")
        print(f"  Summary   : {summary}")

        executions = self.fetch_executions(cycle)
        if not executions:
            print("[WARN] No executions found in this cycle.")
            sys.exit(0)

        print(f"\nFound {len(executions)} execution(s). Fetching test case names...", flush=True)
        keys     = list({ex.get("testCaseKey", "") for ex in executions if ex.get("testCaseKey")})
        tc_names = self.fetch_tc_names(keys)
        print(f"  Fetched {len(tc_names)} test case name(s). Writing Excel...\n")
        self.write_excel(executions, tc_names)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Export test cycle executions to Excel")
    parser.add_argument("--cycle", default="", help="Cycle key (default: from config.properties)")
    parser.add_argument("--out",   default="", help="Output Excel path (.xlsx)")
    args = parser.parse_args()
    CycleExporter(cycle_key=args.cycle, out_path=args.out).run()
