"""
TestCaseUpdater — reads test cases from resources/update_test_case.xlsx
and updates them in Jira Zephyr Scale.

Expected Excel columns:
    Key    : Test case key (required)
    Folder : Folder path (required)
    Status, Owner, Priority, Automation Status,
    Test Reviewer, Test Type, Objective, Precondition, Component, Labels (optional)

Usage:
    python -m test.update_testcases
    python -m test.update_testcases --dry-run
    python -m test.update_testcases --file resources/update_test_case.xlsx
"""

import sys
import os
import argparse

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from main.config import AppConfig
from main.jira_client import JiraClient
from main.report import ReportGenerator


class TestCaseUpdater:

    DEFAULT_XLSX = os.path.join(
        os.path.dirname(__file__), "..", "..", "resources", "update_test_case.xlsx"
    )

    COLUMN_MAP = {
        "key"              : "key",
        "folder"           : "folder",
        "status"           : "status",
        "owner"            : "owner",
        "priority"         : "priority",
        "automation status": "automation_status",
        "test reviewer"    : "test_reviewer",
        "test type"        : "test_type",
        "objective"        : "objective",
        "precondition"     : "precondition",
        "component"        : "component",
        "labels"           : "labels",
        "app mnemonic"     : "app_mnemonic",
    }

    def __init__(self, xlsx_path: str = "", dry_run: bool = False):
        self._cfg       = AppConfig()
        self._client    = JiraClient()
        self.xlsx_path  = xlsx_path or self.DEFAULT_XLSX
        self.dry_run    = dry_run

    # — Read ————————————————————————————————————————————————————————

    def load_xlsx(self) -> list:
        """Read Excel and return list of normalised row dicts."""
        if not os.path.exists(self.xlsx_path):
            print(f"[ERROR] Excel file not found: {os.path.abspath(self.xlsx_path)}")
            sys.exit(1)

        wb = openpyxl.load_workbook(self.xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        raw_headers = [str(cell.value).strip() if cell.value else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        print(f"  Detected columns : {raw_headers}")
        header_lookup = {h.lower(): idx for idx, h in enumerate(raw_headers)}

        if "key" not in header_lookup:
            print(f"[ERROR] 'Key' column is required but not found. Headers: {raw_headers}")
            sys.exit(1)

        if "folder" not in header_lookup:
            print(f"[ERROR] 'Folder' column is required but not found. Headers: {raw_headers}")
            sys.exit(1)

        rows = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            record = {}
            for col_lower, field in self.COLUMN_MAP.items():
                idx = header_lookup.get(col_lower)
                value = str(row[idx]).strip() if idx is not None and row[idx] is not None else ""
                if value == "None":
                    value = ""
                record[field] = value
            if not record.get("key"):
                print(f"  [WARN] Row {i}: 'Key' is empty — skipping.")
                continue
            if not record.get("folder"):
                print(f"  [WARN] Row {i}: 'Folder' is empty for key '{record['key']}' — skipping.")
                continue
            rows.append(record)

        wb.close()
        print(f"  Loaded {len(rows)} test case(s) from: {os.path.basename(self.xlsx_path)}\n")
        return rows

    # — Update ————————————————————————————————————————————————————————

    def build_payload(self, record: dict) -> dict:
        """Build PUT payload from Excel row — only include non-empty fields."""
        cfg = self._cfg
        payload = {}

        if record.get("status"):
            payload["status"] = record["status"]
        # owner: use CSV value if it looks like an account ID, else fall back to config
        owner = record.get("owner") or cfg.owner_account_id
        if owner:
            payload["owner"] = owner
        if record.get("priority"):
            payload["priority"] = record["priority"]
        if record.get("folder"):
            payload["folder"] = record["folder"].strip().rstrip("/")
        if record.get("objective"):
            payload["objective"] = record["objective"]
        if record.get("precondition"):
            payload["precondition"] = record["precondition"]
        if record.get("component"):
            payload["component"] = record["component"]
        if record.get("labels"):
            payload["labels"] = [l.strip() for l in record["labels"].replace(",", " ").split() if l.strip()]

        custom_fields = {}
        # test_reviewer: Excel has display name — use config account ID if set
        test_reviewer = cfg.test_reviewer_id if cfg.test_reviewer_id else record.get("test_reviewer", "")
        if test_reviewer:
            custom_fields["Test Reviewer"] = test_reviewer
        test_type = record.get("test_type") or cfg.default_type_of_test
        if test_type:
            custom_fields["Test Type"] = test_type
        automation_status = record.get("automation_status") or cfg.default_automation_status
        if automation_status:
            custom_fields["Automation Status"] = automation_status
        if record.get("app_mnemonic"):
            custom_fields["App Mnemonic"] = record["app_mnemonic"]
        if custom_fields:
            payload["customFields"] = custom_fields

        return payload

    def update_test_case(self, key: str, payload: dict) -> bool:
        """PUT updated fields onto a single test case. Returns True on success."""
        status, body = self._client.put(f"/testcase/{key}", payload)
        if status not in (200, 204):
            print(f"\n    [DEBUG] {status}: {body}", end="")
        return status in (200, 204)

    # — Run ————————————————————————————————————————————————————————————

    def run(self):
        cfg = self._cfg
        print("=" * 65)
        print("  Jira Zephyr Scale — Update Test Cases from Excel")
        print("=" * 65)
        print(f"  Project  : {cfg.project_key}")
        print(f"  Excel File: {os.path.abspath(self.xlsx_path)}")
        if self.dry_run:
            print("  Mode     : DRY RUN (no changes will be made)")
        print("=" * 65 + "\n")

        records = self.load_xlsx()

        if not records:
            print("[WARN] No valid rows found in Excel file.")
            sys.exit(0)

        ok_list, fail_list = [], []
        test_cases = []

        for i, record in enumerate(records, 1):
            key     = record["key"]
            folder  = record.get("folder", "")
            payload = self.build_payload(record)

            print(f"  [{i:>3}/{len(records)}] {key:20s} ... ", end="", flush=True)

            if self.dry_run:
                print("SKIP (dry-run)")
                test_cases.append({"key": key, "name": folder})
                continue

            if self.update_test_case(key, payload):
                print("✅  Updated")
                ok_list.append(key)
            else:
                print("❌  Failed")
                fail_list.append(key)
            test_cases.append({"key": key, "name": folder})

        print()
        print("=" * 65)
        if self.dry_run:
            print(f"  DRY RUN — {len(records)} test case(s) would be updated.")
        else:
            print(f"  SUMMARY")
            print(f"  Updated : {len(ok_list)}")
            print(f"  Failed  : {len(fail_list)}")
            if ok_list:
                print(f"\n  Updated keys: {', '.join(ok_list)}")
            if fail_list:
                print(f"\n  Failed:")
                for k in fail_list:
                    print(f"    — {k}")
        print("=" * 65)

        if not self.dry_run:
            reports_dir = os.path.join(os.path.dirname(__file__), "..", "..", "reports")
            os.makedirs(reports_dir, exist_ok=True)
            report_path = os.path.join(reports_dir, "Update_TC_Report.html")
            ReportGenerator(output_path=report_path, action_label="Updated", report_title="Zephyr Scale — Update Test Cases Report").generate(
                test_cases=test_cases,
                ok_list=ok_list,
                fail_list=fail_list,
                project_key=cfg.project_key,
                folder_path=records[0].get("folder", "/") if records else "/",
                new_status=records[0].get("status", "") if records else "",
                owner_account_id="",
                test_type="",
            )


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Update test cases from Excel in Jira Zephyr Scale")
    parser.add_argument("--file",    default="", help="Path to Excel file (default: resources/update_test_case.xlsx)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without updating")
    args = parser.parse_args()
    TestCaseUpdater(xlsx_path=args.file, dry_run=args.dry_run).run()
