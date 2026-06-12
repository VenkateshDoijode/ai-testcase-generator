"""
generate_testcases.py — generates test cases using a local Ollama AI model
and writes them to resources/import_test_case.xlsx ready for import.

Requires Ollama running locally: https://ollama.com/download/windows
Recommended model: ollama pull llama3.2

Input options (can combine --issue and --input-folder together):
  --issue       : Jira issue key (fetches description from Jira automatically)
  --input-folder : Folder containing .txt requirement files (reads all .txt files)
  --file        : Single .txt requirements file
  --text        : Requirement text typed inline

Usage:
    python -m test.generate_testcases --issue PROJECT-1234
    python -m test.generate_testcases --input-folder resources/requirements
    python -m test.generate_testcases --issue PROJECT-1234 --input-folder resources/requirements
    python -m test.generate_testcases --file resources/requirements.txt
    python -m test.generate_testcases --text "User should be able to login"
    python -m test.generate_testcases --issue PROJECT-1234 --model mistral --count 10
"""

import sys
import os
import re
import json
import argparse
import urllib.request
import urllib.error
import ssl
import zipfile
import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import docx
import pdfplumber

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from main.config import AppConfig

OLLAMA_BASE_URL    = "http://localhost:11434"
DEFAULT_MODEL      = "google/flan-t5-base"
DEFAULT_COUNT      = 5
HF_ENDPOINT_DEFAULT = "https://your-artifactory-host/artifactory/api/huggingfaceml/huggingface"
OUTPUT_XLSX        = os.path.join(
    os.path.dirname(__file__), "..", "..", "resources", "import_test_case.xlsx"
)

HEADERS = [
    "Folder", "Summary", "Objective", "Status", "Owner",
    "Test Reviewer", "Priority", "Precondition",
    "Test Step", "Test Data", "Expected Result",
    "Traceability", "Type of Test", "Automation Status", "Comments",
]

PROMPT_TEMPLATE = """\
You are a QA engineer. Given the following requirement, generate exactly {count} test cases.

Requirement:
{requirement}

Return ONLY a valid JSON array. Each element must have these exact keys:
  "Summary"        : short test case title
  "Objective"      : what this test case verifies
  "Priority"       : one of High / Medium / Low
  "Precondition"   : preconditions before executing the test
  "Test Step"      : numbered steps to execute (use \\n between steps)
  "Test Data"      : sample input data if applicable
  "Expected Result": what should happen after the steps
  "Type of Test"   : one of Functional / Regression / Negative / Boundary / Smoke
  "Traceability"   : leave blank

Do not include any explanation. Return only the JSON array.
"""

class TestCaseGenerator:

    def __init__(self, model: str = DEFAULT_MODEL, count: int = DEFAULT_COUNT,
                 folder: str = "/Generated", output: str = ""):
        """
        Initialise the generator.

        Args:
            model  : HuggingFace or Ollama model name to use for generation.
            count  : Number of test cases to generate per requirement source.
            folder : Zephyr Scale folder path where test cases will be created.
            output : Full path to the output Excel file (defaults to resources/import_test_case.xlsx).
        """
        self._cfg    = AppConfig()
        self.model   = model
        self.count   = count
        self.folder  = folder
        self.output  = output or OUTPUT_XLSX
        self._ssl    = ssl.create_default_context()

    # — Input sources ————————————————————————————————————————————

    def fetch_jira_issue(self, issue_key: str) -> str:
        """
        Fetch the summary and description of a Jira issue via the REST API.

        Args:
            issue_key : Jira issue key (e.g. PROJECT-1234).

        Returns:
            Combined summary and description as a plain-text string.
        """
        cfg = self._cfg
        url = f"{cfg.base_url}/rest/api/2/issue/{issue_key}?fields=summary,description"
        req = urllib.request.Request(url, headers={
            "Authorization": f"Bearer {cfg.jira_token}",
            "Accept": "application/json",
        })
        try:
            with urllib.request.urlopen(req, context=self._ssl) as resp:
                data   = json.loads(resp.read().decode())
                fields = data.get("fields", {})
                summary = fields.get("summary", "")
                desc    = fields.get("description", "") or ""
                if isinstance(desc, dict):
                    desc = self._extract_adf_text(desc)
                text = f"Summary: {summary}\n\nDescription:\n{desc}".strip()
                print(f"  Fetched issue: {issue_key} — {summary}")
                return text
        except urllib.error.HTTPError as e:
            print(f"[ERROR] Could not fetch issue {issue_key}: {e.code}")
            sys.exit(1)

    def fetch_confluence_page(self, page_input: str) -> tuple:
        """
        Fetch plain-text content from a Confluence page.

        Args:
            page_input : Confluence page ID (numeric) or full page URL.

        Returns:
            Tuple of (page_title, plain_text_content).
        """
        cfg = self._cfg
        page_id = page_input.strip()
        if page_input.startswith("http"):
            m = re.search(r'pageId=(\d+)|/pages/(\d+)', page_input)
            if not m:
                print(f"[ERROR] Could not extract page ID from URL: {page_input}")
                sys.exit(1)
            page_id = m.group(1) or m.group(2)
        base = cfg.base_url.rstrip("/")
        url  = f"{base}/wiki/rest/api/content/{page_id}?expand=body.storage,title"
        req = urllib.request.Request(url, headers={
            "Authorization": f"Bearer {cfg.jira_token}",
            "Accept": "application/json",
        })
        try:
            with urllib.request.urlopen(req, context=self._ssl) as resp:
                data  = json.loads(resp.read().decode())
                title = data.get("title", page_id)
                html  = data.get("body", {}).get("storage", {}).get("value", "")
                text  = self._strip_html(html)
                print(f"  Fetched Confluence page: {title} ({len(text)} chars)")
                return title, text
        except urllib.error.HTTPError as e:
            print(f"[ERROR] Could not fetch Confluence page {page_id}: {e.code} {e.reason}")
            sys.exit(1)

    def _strip_html(self, html: str) -> str:
        """
        Strip HTML/XML tags from a string and decode common HTML entities.

        Args:
            html : Raw HTML string from Confluence storage format.

        Returns:
            Clean plain-text string.
        """
        text = re.sub(r'<[^>]+>', ' ', html)
        text = text.replace('&nbsp;', ' ').replace('&amp;', '&') \
                    .replace('&lt;', '<').replace('&gt;', '>') \
                    .replace('&quot;', '"').replace('&#39;', "'")
        text = re.sub(r'[ \t]+', ' ', text)
        text = re.sub(r'\n{3,}', '\n\n', text)
        return text.strip()

    def _extract_adf_text(self, node: dict) -> str:
        """
        Recursively extract plain text from an Atlassian Document Format (ADF) node.

        Args:
            node : ADF node dict as returned by the Jira REST API description field.

        Returns:
            Plain-text string with all nested text nodes joined.
        """
        if not isinstance(node, dict):
            return str(node)
        if node.get("type") == "text":
            return node.get("text", "")
        parts = []
        for child in node.get("content", []):
            parts.append(self._extract_adf_text(child))
        return " ".join(p for p in parts if p)

    def load_text_file(self, path: str) -> str:
        """
        Load and return the contents of a single plain-text (.txt) requirement file.

        Args:
            path : Absolute or relative path to the .txt file.

        Returns:
            File contents as a stripped string.
        """
        if not os.path.exists(path):
            print(f"[ERROR] File not found: {path}")
            sys.exit(1)
        with open(path, encoding="utf-8") as f:
            return f.read().strip()

    def load_folder(self, folder: str) -> list:
        """
        Load all requirement files (.txt, .docx, .pdf) from a folder.

        Args:
            folder : Path to the folder containing requirement files.

        Returns:
            List of (filename, text_content) tuples for each successfully read file.
        """
        if not os.path.isdir(folder):
            print(f"[ERROR] Folder not found: {folder}")
            sys.exit(1)
        files = sorted(f for f in os.listdir(folder) if f.endswith(".txt") or f.endswith(".docx") or f.endswith(".pdf"))
        if not files:
            print(f"[WARN] No .txt, .docx or .pdf files found in: {folder}")
            return []
        results = []
        for fname in files:
            path = os.path.join(folder, fname)
            if fname.endswith(".docx"):
                text = self._read_docx(path)
            elif fname.endswith(".pdf"):
                text = self._read_pdf(path)
            else:
                try:
                    with open(path, encoding="utf-8") as f:
                        text = f.read().strip()
                except UnicodeDecodeError:
                    with open(path, encoding="latin-1") as f:
                        text = f.read().strip()
                    print(f"  [WARN] {fname}: UTF-8 decode failed, read as latin-1")
            if text:
                results.append((fname, text))
                print(f"  Loaded: {fname} ({len(text)} chars)")
        return results

    def _read_docx(self, path: str) -> str:
        """
        Extract plain text from a Word .docx file.
        Tries python-docx first, falls back to raw XML parsing if that fails.

        Args:
            path : Path to the .docx file.

        Returns:
            Extracted plain-text string, or empty string on failure.
        """
        try:
            doc   = docx.Document(path)
            lines = [para.text.strip() for para in doc.paragraphs if para.text.strip()]
            return "\n".join(lines)
        except Exception:
            pass
        try:
            with zipfile.ZipFile(path, "r") as z:
                with z.open("word/document.xml") as f:
                    tree = ET.parse(f)
                    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
                    texts = [t.text for t in tree.findall(".//w:t", ns) if t.text]
                    return "\n".join(texts)
        except Exception as e:
            print(f"  [WARN] Could not read {os.path.basename(path)}: {e}")
            return ""

    def _read_pdf(self, path: str) -> str:
        """
        Extract plain text from a PDF file using pdfplumber.

        Args:
            path : Path to the .pdf file.

        Returns:
            Extracted plain-text string with all pages joined by newlines.
        """
        lines = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    lines.append(text.strip())
        return "\n".join(lines)

    # — Ollama ————————————————————————————————————————————

    def _ollama_running(self) -> bool:
        """
        Check whether a local Ollama server is running and reachable.

        Returns:
            True if Ollama is running at OLLAMA_BASE_URL, False otherwise.
        """
        try:
            urllib.request.urlopen(f"{OLLAMA_BASE_URL}/api/tags", timeout=3)
            return True
        except Exception:
            return False

    def _load_hf_pipeline(self):
        """
        Load and cache a HuggingFace transformers text2text-generation pipeline.
        Sets HF_ENDPOINT and HF_TOKEN environment variables from config.ini before loading.

        Returns:
            Loaded pipeline object, or None if loading fails.
        """
        if getattr(self, '_pipeline', None):
            return self._pipeline
        try:
            import os
            hf_endpoint = self._cfg.hf_endpoint or HF_ENDPOINT_DEFAULT
            hf_token    = self._cfg.hf_token
            os.environ['HF_ENDPOINT']             = hf_endpoint
            os.environ['HF_HUB_ETAG_TIMEOUT']     = '86400'
            os.environ['HF_HUB_DOWNLOAD_TIMEOUT'] = '86400'
            if hf_token:
                os.environ['HF_TOKEN'] = hf_token
            from transformers import pipeline
            print(f"  Loading HuggingFace model: {self.model} (may download on first run)...")
            self._pipeline = pipeline("text2text-generation", model=self.model)
            print("  Model loaded.")
            return self._pipeline
        except Exception as e:
            print(f"  [WARN] HuggingFace model load failed: {e})")
            self._pipeline = None
            return None

    def generate(self, requirement: str) -> list:
        """
        Generate test cases from a requirement string.
        Tries engines in order: HuggingFace → Ollama → rule-based fallback.

        Args:
            requirement : Plain-text requirement or user story.

        Returns:
            List of test case dicts with keys: Summary, Objective, Priority,
            Precondition, Test Step, Test Data, Expected Result, Type of Test, Traceability.
        """
        pipe = self._load_hf_pipeline()
        if pipe:
            return self._generate_hf(requirement, pipe)
        if self._ollama_running():
            print("  [INFO] Using Ollama as fallback.")
            return self._generate_ollama(requirement)
        print("  [INFO] No AI engine available — using rule-based generator.")
        return self._generate_rulebased(requirement)

    def _generate_hf(self, requirement: str, pipe) -> list:
        """
        Generate test cases using a loaded HuggingFace transformers pipeline.
        Falls back to rule-based generation if the pipeline call fails.

        Args:
            requirement : Plain-text requirement string.
            pipe        : Loaded HuggingFace pipeline object.

        Returns:
            List of test case dicts.
        """
        print(f"  Generating with HuggingFace ({self.model})...")
        prompt = PROMPT_TEMPLATE.format(count=self.count, requirement=requirement)
        try:
            out      = pipe(prompt, max_new_tokens=512, do_sample=False)
            raw_text = out[0]["generated_text"].strip() if out else ""
            return self._parse_json_response(raw_text)
        except Exception as e:
            print(f"  [WARN] HuggingFace generation failed: {e} — falling back to rule-based.")
            return self._generate_rulebased(requirement)

    def _generate_ollama(self, requirement: str) -> list:
        """
        Generate test cases by calling the local Ollama REST API.
        Falls back to rule-based generation on network failure.

        Args:
            requirement : Plain-text requirement string.

        Returns:
            List of test case dicts.
        """
        prompt  = PROMPT_TEMPLATE.format(count=self.count, requirement=requirement)
        payload = json.dumps({"model": self.model, "prompt": prompt, "stream": False}).encode("utf-8")
        req = urllib.request.Request(
            f"{OLLAMA_BASE_URL}/api/generate",
            data=payload, headers={"Content-Type": "application/json"}, method="POST",
        )
        print(f"  Sending to Ollama ({self.model}) — please wait...")
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                raw_text = json.loads(resp.read().decode()).get("response", "").strip()
        except urllib.error.URLError as e:
            print(f"  [WARN] Ollama request failed: {e} — falling back to rule-based.")
            return self._generate_rulebased(requirement)
        return self._parse_json_response(raw_text)

    def _generate_rulebased(self, requirement: str) -> list:
        """
        Generate test cases using a deterministic rule-based engine.
        Used as the final fallback when no AI engine is available.
        Cycles through Positive, Negative, Boundary, Smoke, Regression, Security,
        and Performance test patterns.

        Args:
            requirement : Plain-text requirement string.

        Returns:
            List of test case dicts.
        """
        print("  Generating with rule-based engine...")
        lines = [l.strip() for l in requirement.split("\n") if len(l.strip()) > 15]
        if not lines:
            lines = [requirement[:100]]
        keywords = [
            ("Positive",   "Verify {topic} with valid input",        "Functional", "High"),
            ("Negative",   "Verify {topic} with invalid input",       "Functional", "High"),
            ("Boundary",   "Verify {topic} at boundary values",       "Boundary",   "Medium"),
            ("Empty input","Verify {topic} with empty/null input",    "Negative",   "Medium"),
            ("Smoke",      "Smoke test for {topic}",                  "Smoke",      "High"),
            ("Regression", "Regression test for {topic}",             "Regression", "Medium"),
            ("Security",   "Verify {topic} access control",           "Functional", "Low"),
            ("Performance","Verify {topic} response time",            "Functional", "Low"),
        ]
        test_cases = []
        for i in range(self.count):
            line  = lines[i % len(lines)]
            topic = line[:70]
            ktype = keywords[i % len(keywords)]
            test_cases.append({
                "Summary"        : ktype[1].format(topic=topic),
                "Objective"      : f"{ktype[0]} test — {topic}",
                "Priority"       : ktype[3],
                "Precondition"   : "System is running and user is authenticated",
                "Test Step"      : f"1. Navigate to the relevant module\n2. Perform action: {topic}\n3. Observe the result",
                "Test Data"      : "As per requirement",
                "Expected Result": f"System behaves as expected for: {topic}",
                "Type of Test"   : ktype[2],
                "Traceability"   : "",
            })
        return test_cases

    def _parse_json_response(self, text: str) -> list:
        """
        Parse and extract a JSON array from a raw AI model response string.
        Locates the first '[' and last ']' to isolate the array even if
        the model returns extra prose around it.

        Args:
            text : Raw text response from the AI model.

        Returns:
            Parsed list of test case dicts.

        Raises:
            SystemExit if no valid JSON array is found.
        """
        start = text.find("[")
        end   = text.rfind("]") + 1
        if start == -1 or end == 0:
            print("[ERROR] Model did not return a valid JSON array.")
            print(f"  Raw response:\n{text[:500]}")
            sys.exit(1)
        try:
            return json.loads(text[start:end])
        except json.JSONDecodeError as e:
            print(f"[ERROR] Failed to parse JSON from model response: {e}")
            print(f"  Raw:\n{text[:500]}")
            sys.exit(1)

    # — Write Excel ————————————————————————————————————————————

    def write_xlsx(self, test_cases: list, traceability: str = "", append: bool = False):
        """
        Write generated test cases to the output Excel file.
        Creates a new file or appends to an existing one based on the append flag.

        Args:
            test_cases   : List of test case dicts to write.
            traceability : Optional Jira issue key to set as traceability for all rows.
            append       : If True and the output file exists, append rows instead of overwriting.
        """
        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill(fill_type="solid", fgColor="0052CC")
        header_align = Alignment(horizontal="center", vertical="center")

        if append and os.path.exists(self.output):
            wb = openpyxl.load_workbook(self.output)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Generated Test Cases"
            for col_idx, col_name in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.font      = header_font
                cell.fill      = header_fill
                cell.alignment = header_align

        for tc in test_cases:
            ws.append([
                self.folder,
                tc.get("Summary", ""),
                tc.get("Objective", ""),
                "",
                "",
                "",
                tc.get("Priority", "Medium"),
                tc.get("Precondition", ""),
                tc.get("Test Step", ""),
                tc.get("Test Data", ""),
                tc.get("Expected Result", ""),
                traceability or tc.get("Traceability", ""),
                tc.get("Type of Test", "Functional"),
                "Manual",
                "",
            ])

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        os.makedirs(os.path.dirname(os.path.abspath(self.output)), exist_ok=True)
        wb.save(self.output)
        print(f"\n  ✅ {len(test_cases)} test case(s) written to: {os.path.abspath(self.output)}")
        print("  Run import:  python -m test.import_testcases")

    # — Run ————————————————————————————————————————————

    def run(self, issue: str = "", text: str = "", file: str = "", input_folder: str = "", confluence: str = ""):
        """
        Main entry point — collects requirements from all specified sources,
        generates test cases for each, and writes them to the output Excel file.

        Args:
            issue        : Jira issue key to fetch requirement from.
            text         : Inline requirement text.
            file         : Path to a single .txt requirement file.
            input_folder : Path to a folder containing requirement files.
            confluence   : Confluence page ID or URL (comma-separated for multiple).
        """
        print("=" * 65)
        print("  Jira Zephyr Scale — AI Test Case Generator (Ollama)")
        print("=" * 65)
        print(f"  Model  : {self.model}")
        print(f"  Count  : {self.count}")
        print(f"  Folder : {self.folder}")
        print(f"  HF URL : {self._cfg.hf_endpoint or HF_ENDPOINT_DEFAULT}")
        print("=" * 65 + "\n")

        sources = []  # list of (requirement_text, traceability)

        if issue:
            req_text = self.fetch_jira_issue(issue)
            sources.append((req_text, issue))

        if confluence:
            for page_ref in [p.strip() for p in confluence.split(",") if p.strip()]:
                title, req_text = self.fetch_confluence_page(page_ref)
                sources.append((req_text, title))

        if input_folder:
            files = self.load_folder(input_folder)
            for fname, req_text in files:
                print(f"  Loaded file: {fname}")
                sources.append((req_text, ""))

        if file:
            req_text = self.load_text_file(file)
            sources.append((req_text, ""))

        if text:
            sources.append((text.strip(), ""))

        if not sources:
            print("[ERROR] Provide at least one of: --issue, --input-folder, --file, --text, --confluence.")
            sys.exit(1)

        all_test_cases = []
        total_written  = 0

        for idx, (requirement, traceability) in enumerate(sources):
            if not requirement:
                continue
            label = traceability or f"source {idx + 1}"
            print(f"\n --- Generating for: {label} ---")
            print(f"  Requirement preview: {requirement[:120]}...\n")

            test_cases = self.generate(requirement)
            print(f"  Generated {len(test_cases)} test case(s)")
            for i, tc in enumerate(test_cases, 1):
                print(f"    [{i}] {tc.get('Summary', '(no summary)')}")

            append = (idx > 0) or (total_written > 0)
            self.write_xlsx(test_cases, traceability, append=append)
            total_written += len(test_cases)

        print(f"\n  Total test cases written: {total_written}")
        print("  Run import:  python -m test.import_testcases")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Generate test cases using Ollama AI and export to Excel"
    )
    parser.add_argument('--issue',        default="", help="Jira issue key (e.g. PROJECT-1234)")
    parser.add_argument('--input-folder', default="", dest='input_folder', help="Folder with .txt requirement files")
    parser.add_argument('--file',         default="", help="Single .txt requirements file")
    parser.add_argument('--text',         default="", help="Requirement text inline")
    parser.add_argument('--confluence',   default="", help="Confluence page ID or URL (comma-separated for multiple)")
    parser.add_argument('--model',        default=DEFAULT_MODEL, help=f"Ollama model (default: {DEFAULT_MODEL})")
    parser.add_argument('--count',        type=int, default=DEFAULT_COUNT, help=f"Test cases per source (default: {DEFAULT_COUNT})")
    parser.add_argument('--folder',       default="/Generated", help="Zephyr folder path (default: /Generated)")
    parser.add_argument('--output',       default="", help="Output Excel path (default: resources/import_test_case.xlsx)")

    args = parser.parse_args()
    if not any([args.issue, args.input_folder, args.file, args.text, args.confluence]):
        parser.error("Provide at least one of: --issue, --input-folder, --file, --text, --confluence")

    TestCaseGenerator(
        model=args.model,
        count=args.count,
        folder=args.folder,
        output=args.output,
    ).run(issue=args.issue, text=args.text, file=args.file, input_folder=args.input_folder, confluence=args.confluence)
